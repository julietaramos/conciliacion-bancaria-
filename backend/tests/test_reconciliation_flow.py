"""
Cubre los 6 cambios pedidos sobre el flujo de conciliación:
  1) margen de $100 al tildar / en la diferencia final
  2) cruces mezclando cualquiera de las 4 columnas (col1..col4)
  3-4-5) no tocan al backend (orden de pasos, UI de totales, destildar)
  6) no toca al backend (orden de tarjetas de banco, solo localStorage)
"""
import io
from datetime import datetime, timezone

import pytest
from openpyxl import load_workbook

from db import crud
from main import _match_to_internal, _result_from_state
from utils.excel_report import generate_excel_report


def _item(desc, monto, mes_anterior=False, fecha="2026-03-01T00:00:00"):
    return {"fecha": fecha, "descripcion": desc, "monto": monto, "mes_anterior": mes_anterior}


# ── 1) Margen de $100 ───────────────────────────────────────────────────────────

@pytest.mark.parametrize("diferencia,estado_esperado", [
    (0,      "balanceada"),
    (50,     "balanceada"),
    (-99.99, "balanceada"),   # el margen es simétrico
    (99.99,  "balanceada"),
    (100,    "con_diferencias"),   # el límite es estrictamente menor a 100
    (100.01, "con_diferencias"),
    (500,    "con_diferencias"),
])
def test_tolerancia_100_pesos_en_estado(db_session, diferencia, estado_esperado):
    result = {"diferencia": diferencia, "col1": [], "col2": [], "col3": [], "col4": [], "matched": []}
    row = crud.save_conciliacion(db_session, result, banco_id=None)
    assert row.estado == estado_esperado


# ── 2) Cruces mezclando las 4 columnas ─────────────────────────────────────────

def test_match_to_internal_soporta_las_4_columnas_en_un_solo_cruce():
    m = {
        "id": "m1",
        "col1": [_item("debito A", 100)],
        "col2": [_item("haber A", 60)],
        "col3": [],
        "col4": [_item("credito A", 40)],
    }
    internal = _match_to_internal(m)
    assert internal["id"] == "m1"
    assert [x["monto"] for x in internal["col1"]] == [100]
    assert [x["monto"] for x in internal["col2"]] == [60]
    assert internal["col3"] == []
    assert [x["monto"] for x in internal["col4"]] == [40]
    # fecha debe quedar parseada a datetime, no como string
    assert isinstance(internal["col1"][0]["fecha"], datetime)


def test_result_from_state_calcula_partidas_con_cruce_mixto_ya_descontado():
    # Los ítems cruzados NO deben aparecer en col1..col4 (partidas pendientes) —
    # el frontend ya los sacó de esas listas al cruzar; acá solo llegan los que
    # quedaron sin cruzar más el/los matched.
    state = {
        "col1": [_item("debito pendiente", 30)],
        "col2": [],
        "col3": [],
        "col4": [_item("credito pendiente", 40)],
        "matched": [{
            "id": "m1",
            "col1": [_item("debito A", 100)],
            "col2": [_item("haber A", 60)],
            "col3": [],
            "col4": [_item("credito A", 40)],
        }],
        "ajustes": [],
        "saldo_banco": 1000,
        "saldo_contable": 950,
    }
    result = _result_from_state(state)
    # partidas = col1 - col2 + col3 - col4 = 30 - 0 + 0 - 40 = -10
    assert result["partidas"] == -10
    assert len(result["matched"]) == 1
    assert result["matched"][0]["col1"][0]["monto"] == 100


# ── Reporte Excel: tabla de cruces confirmados con 4 columnas ─────────────────

def _cell_values(ws, row):
    return tuple(ws.cell(row=row, column=c).value for c in range(1, 14))


def test_excel_report_tabla_cruces_mixtos():
    result = {
        "saldo_banco": 1000.0, "saldo_contable": 950.0, "partidas": 50.0, "diferencia": 0.0,
        "col1": [_item("debito pendiente", 30)],
        "col2": [_item("haber pendiente", 20)],
        "col3": [],
        "col4": [_item("credito pendiente", 40)],
        "matched": [
            # cruce mixto: débito + haber + crédito juntos
            {"id": "m1", "col1": [_item("debito A", 100)], "col2": [_item("haber A", 60)],
             "col3": [], "col4": [_item("credito A", 40)]},
            # cruce clásico 1:1 (estilo pre-cambio), con marca de mes anterior
            {"id": "m2", "col1": [], "col2": [_item("haber B", 15, mes_anterior=True)],
             "col3": [_item("debe B", 15)], "col4": []},
        ],
        "ajustes": [],
    }
    data = generate_excel_report(result)
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active

    # Fila del cruce mixto: débito=100 (col A-C), haber=60 (col D-F), debe vacío (H-J), crédito=40 (K-M)
    row = _cell_values(ws, 15)
    assert row[1] == "debito A" and row[2] == 100
    assert row[4] == "haber A" and row[5] == 60
    assert row[7] is None and row[8] is None and row[9] is None
    assert row[11] == "credito A" and row[12] == 40

    # Fila del cruce clásico: solo haber (D-F) y debe (H-J)
    row = _cell_values(ws, 16)
    assert row[0] is None
    assert row[4] == "haber B" and row[5] == 15
    assert row[8] == "debe B" and row[9] == 15

    # Fila TOTAL: suma por columna a través de todos los cruces
    total = _cell_values(ws, 17)
    assert total[2] == 100          # col1
    assert total[5] == 75           # col2 = 60 + 15
    assert total[9] == 15           # col3
    assert total[12] == 40          # col4


def test_excel_report_sin_cruces_no_rompe():
    result = {
        "saldo_banco": 100.0, "saldo_contable": 100.0, "partidas": 0.0, "diferencia": 0.0,
        "col1": [], "col2": [], "col3": [], "col4": [],
        "matched": [],
        "ajustes": [],
    }
    data = generate_excel_report(result)
    assert len(data) > 0


# ── Integración end-to-end vía el endpoint real ───────────────────────────────

def test_endpoint_generar_guarda_cruces_mixtos_y_respeta_margen(client):
    banco = client.post("/api/bancos", json={"nombre": "Banco Test"}).json()
    banco_id = banco["id"]

    body = {
        "banco_id": banco_id,
        "col1": [_item("debito pendiente", 30)],
        "col2": [],
        "col3": [],
        "col4": [],
        "matched": [{
            "id": "m1",
            "col1": [_item("debito A", 100)],
            "col2": [_item("haber A", 60)],
            "col3": [],
            "col4": [_item("credito A", 40)],
        }],
        "ajustes": [],
        "saldo_banco": 1030,     # 1000 + 30 pendiente
        "saldo_contable": 1000,  # sin ajustes
        "fecha_datos": "2026-03-01",
    }
    res = client.post("/api/conciliar/generar", json=body)
    assert res.status_code == 200
    assert res.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert len(res.content) > 0

    listado = client.get("/api/bancos").json()
    ult = next(b for b in listado if b["id"] == banco_id)["ultima_conciliacion"]
    # partidas = col1(30) = 30 -> diferencia = saldo_banco(1030) + 30 - saldo_contable(1000) = 60
    assert ult["diferencia"] == 60.0
    assert ult["estado"] == "balanceada"       # 60 < 100, dentro del margen nuevo
    assert ult["resumen"]["cruces_confirmados"] == 1
    assert ult["resumen"]["debitos_no_contab"] == 1


def test_endpoint_generar_fuera_de_margen_queda_con_diferencias(client):
    banco = client.post("/api/bancos", json={"nombre": "Banco Test 2"}).json()
    banco_id = banco["id"]

    body = {
        "banco_id": banco_id,
        "col1": [], "col2": [], "col3": [], "col4": [],
        "matched": [],
        "ajustes": [],
        "saldo_banco": 1200,
        "saldo_contable": 1000,   # diferencia = 200, fuera del margen de 100
        "fecha_datos": "2026-03-01",
    }
    res = client.post("/api/conciliar/generar", json=body)
    assert res.status_code == 200

    listado = client.get("/api/bancos").json()
    ult = next(b for b in listado if b["id"] == banco_id)["ultima_conciliacion"]
    assert ult["diferencia"] == 200.0
    assert ult["estado"] == "con_diferencias"
