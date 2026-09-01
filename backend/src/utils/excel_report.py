import io
from datetime import date as dt_date
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

_TABLE_STYLE = TableStyleInfo(
    name="TableStyleLight1", showRowStripes=False, showColumnStripes=False,
    showFirstColumn=False, showLastColumn=False,
)


def _add_table(ws, name: str, ref: str):
    """Attach an independent AutoFilter (Excel Table) scoped to `ref` only —
    filtering it never hides rows belonging to other tables on the sheet."""
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = _TABLE_STYLE
    ws.add_table(table)

# ── Palette ────────────────────────────────────────────────────────────────────
_DARK_BLUE  = "0D1B4B"
_MED_BLUE   = "1A3C8F"
_LIGHT_BLUE = "C7D0E8"
_YELLOW     = "FFF3CD"
_GREEN      = "D4EDDA"
_RED        = "F8D7DA"
_PURPLE     = "E8D5F5"
_GREY_LIGHT = "F5F5F5"
_WHITE      = "FFFFFF"

_THIN = Side(style="thin", color="AAAAAA")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_NUM_FMT = '#,##0.00'
_DATE_FMT = 'DD/MM/YYYY'


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=10) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri")


def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _set_header_row(ws, row: int, cells: dict, bg: str, fg: str = "FFFFFF", size=10):
    for col, text in cells.items():
        c = ws.cell(row=row, column=col)
        c.value = text
        c.fill = _fill(bg)
        c.font = _font(bold=True, color=fg, size=size)
        c.alignment = _align("center", wrap=True)
        c.border = _BORDER


def _write_data_row(ws, row: int, values: list, bg: str = None):
    for col, val in enumerate(values, start=1):
        c = ws.cell(row=row, column=col, value=val)
        c.border = _BORDER
        if bg:
            c.fill = _fill(bg)
        if isinstance(val, float):
            c.number_format = _NUM_FMT
            c.alignment = _align("right")
        elif hasattr(val, "date") or isinstance(val, dt_date):
            c.number_format = _DATE_FMT
            c.alignment = _align("center")
        else:
            c.alignment = _align("left", wrap=True)


def _fmt_date(item: dict):
    f = item.get("fecha")
    if f is None:
        return None
    if hasattr(f, "to_pydatetime"):
        return f.to_pydatetime()
    return f


# ── Main generator ─────────────────────────────────────────────────────────────

def generate_excel_report(result: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Conciliacion"

    col1 = result["col1"]   # extracto DÉBITO  unmatched
    col2 = result["col2"]   # mayor   HABER    unmatched
    col3 = result["col3"]   # mayor   DEBE     unmatched
    col4 = result["col4"]   # extracto CRÉDITO unmatched
    matched = result["matched"]  # cruces confirmados, cada uno con ítems de cualquiera de las 4 columnas

    # ── Column widths ───────────────────────────────────────────────────────────
    widths = {
        1: 14, 2: 50, 3: 16,          # col group 1
        4: 14, 5: 50, 6: 16,          # col group 2
        7: 3,                          # separator
        8: 14, 9: 50, 10: 16,         # col group 3
        11: 14, 12: 50, 13: 16,       # col group 4
    }
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # ── Rows 1-5: Summary block ─────────────────────────────────────────────────
    summary = [
        ("Saldo Banco (Final)",        result["saldo_banco"],    "A"),
        ("Partidas Pendientes",         result["partidas"],       "B"),
        ("Saldo Banco + Partidas",      result["saldo_banco"] + result["partidas"], "C = A + B"),
        ("Saldo Contable (Final Mayor)", result["saldo_contable"], "D"),
        ("Diferencia",                  result["diferencia"],     "C - D"),
    ]
    for i, (label, value, ref) in enumerate(summary, start=1):
        ws.cell(row=i, column=6, value=label).font = _font(bold=(i == 5))
        c_val = ws.cell(row=i, column=8, value=value)
        c_val.number_format = _NUM_FMT
        c_val.font = _font(bold=True, color=("C0392B" if i == 5 and abs(value) > 0.01 else "155724"))
        c_val.alignment = _align("right")
        ws.cell(row=i, column=9, value=ref).font = _font(color="888888")

    # ── Row 6: Group headers ────────────────────────────────────────────────────
    sum1 = sum(x["monto"] for x in col1)
    sum2 = sum(x["monto"] for x in col2)
    sum3 = sum(x["monto"] for x in col3)
    sum4 = sum(x["monto"] for x in col4)

    group_headers = {
        1: "Débitos no Contabilizados",
        3: sum1,
        4: "No Debitados en Extracto",
        6: -sum2,
        8: "No Acreditados",
        10: sum3,
        11: "Créditos no Contabilizados",
        13: -sum4,
    }
    _set_header_row(ws, 6, group_headers, _DARK_BLUE)
    ws.cell(row=6, column=3).number_format = _NUM_FMT
    ws.cell(row=6, column=6).number_format = _NUM_FMT
    ws.cell(row=6, column=10).number_format = _NUM_FMT
    ws.cell(row=6, column=13).number_format = _NUM_FMT
    ws.row_dimensions[6].height = 36

    # ── Row 7: Sub-headers ──────────────────────────────────────────────────────
    sub = {1: "Fecha", 2: "Concepto", 3: "Importe",
           4: "Fecha", 5: "Concepto", 6: "Importe",
           8: "Fecha", 9: "Concepto", 10: "Importe",
           11: "Fecha", 12: "Concepto", 13: "Importe"}
    _set_header_row(ws, 7, sub, _MED_BLUE)

    # ── Rows 8+: Data ───────────────────────────────────────────────────────────
    max_rows = max(len(col1), len(col2), len(col3), len(col4), 1)

    def _item(lst, i):
        return lst[i] if i < len(lst) else None

    for i in range(max_rows):
        row = 8 + i

        a = _item(col1, i)
        b = _item(col2, i)
        c = _item(col3, i)
        d = _item(col4, i)

        # If any item in this row comes from the previous month, tint orange
        any_anterior = any(x.get("mes_anterior") for x in [a, b, c, d] if x)
        bg = _ORANGE if any_anterior else (_GREY_LIGHT if i % 2 == 0 else _WHITE)

        row_vals = [
            _fmt_date(a) if a else None, a["descripcion"] if a else None, a["monto"] if a else None,
            _fmt_date(b) if b else None, b["descripcion"] if b else None, b["monto"] if b else None,
            None,
            _fmt_date(c) if c else None, c["descripcion"] if c else None, c["monto"] if c else None,
            _fmt_date(d) if d else None, d["descripcion"] if d else None, d["monto"] if d else None,
        ]
        _write_data_row(ws, row, row_vals, bg)

    # "Arriba" (no tildado): partidas pendientes — cada columna de datos con su
    # propia tabla/filtro, independiente de las demás y de los conciliados de abajo.
    last_pending_row = 7 + max_rows
    _add_table(ws, "DebitosNoContabilizados",  f"A7:C{last_pending_row}")
    _add_table(ws, "NoDebitadosEnExtracto",    f"D7:F{last_pending_row}")
    _add_table(ws, "NoAcreditados",            f"H7:J{last_pending_row}")
    _add_table(ws, "CreditosNoContabilizados", f"K7:M{last_pending_row}")

    # ── Verification table ──────────────────────────────────────────────────────
    # "Abajo" (tildado): cruces conciliados — cada uno puede combinar ítems de
    # cualquiera de las 4 columnas, así que se listan en el mismo layout de 4
    # grupos que la sección de pendientes, una tabla por grupo.
    start_row = 8 + max_rows + 3
    start_row = _write_matched_table(ws, start_row, matched)

    # ── Ajustes de saldo contable (errores de meses anteriores) ────────────────
    ajustes = result.get("ajustes") or []
    if ajustes:
        start_row += 2
        _write_ajustes_section(
            ws, start_row, ajustes,
            saldo_original=result.get("saldo_contable_original", result["saldo_contable"]),
            saldo_ajustado=result["saldo_contable"],
        )

    # ── Freeze top rows ─────────────────────────────────────────────────────────
    ws.freeze_panes = "A8"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


_ORANGE = "FFE0B2"  # highlight for previous-month items


def _write_matched_table(ws, start_row: int, matched: list) -> int:
    """Write the confirmed-crosses table. A cross can combine items from any
    of the 4 columns (not just a fixed Mayor/Extracto pair), so it's laid out
    with the same 4-group shape as the pending section above. Returns the
    next available row."""

    c = ws.cell(row=start_row, column=1, value="✓ Cruces Confirmados")
    c.font = _font(bold=True, size=11, color=_DARK_BLUE)
    c.fill = _fill(_YELLOW)
    for col in range(2, 14):
        ws.cell(row=start_row, column=col).fill = _fill(_YELLOW)
    start_row += 1

    group_row = start_row
    group_labels = {
        1: "Débito (Extracto)", 4: "Haber (Mayor)",
        8: "Debe (Mayor)", 11: "Crédito (Extracto)",
    }
    _set_header_row(ws, group_row, group_labels, _DARK_BLUE, size=9)
    start_row += 1

    header_row = start_row
    sub = {1: "Fecha", 2: "Concepto", 3: "Importe",
           4: "Fecha", 5: "Concepto", 6: "Importe",
           8: "Fecha", 9: "Concepto", 10: "Importe",
           11: "Fecha", 12: "Concepto", 13: "Importe"}
    _set_header_row(ws, header_row, sub, _MED_BLUE, size=9)
    start_row += 1

    if not matched:
        ws.cell(row=start_row, column=1, value="— Sin cruces confirmados —").font = _font(color="888888")
        return start_row + 1

    def _item(lst, i):
        return lst[i] if i < len(lst) else None

    for i, m in enumerate(matched):
        c1, c2, c3, c4 = m.get("col1", []), m.get("col2", []), m.get("col3", []), m.get("col4", [])
        any_anterior = any(x.get("mes_anterior") for grp in (c1, c2, c3, c4) for x in grp)
        bg = _ORANGE if any_anterior else (_GREY_LIGHT if i % 2 == 0 else _WHITE)
        max_j = max(len(c1), len(c2), len(c3), len(c4), 1)
        for j in range(max_j):
            a, b, cc, d = _item(c1, j), _item(c2, j), _item(c3, j), _item(c4, j)
            row_vals = [
                _fmt_date(a) if a else None, a["descripcion"] if a else None, a["monto"] if a else None,
                _fmt_date(b) if b else None, b["descripcion"] if b else None, b["monto"] if b else None,
                None,
                _fmt_date(cc) if cc else None, cc["descripcion"] if cc else None, cc["monto"] if cc else None,
                _fmt_date(d) if d else None, d["descripcion"] if d else None, d["monto"] if d else None,
            ]
            _write_data_row(ws, start_row, row_vals, bg)
            start_row += 1

    last_data_row = start_row - 1

    totals = {
        3:  sum(x["monto"] for m in matched for x in m.get("col1", [])),
        6:  sum(x["monto"] for m in matched for x in m.get("col2", [])),
        10: sum(x["monto"] for m in matched for x in m.get("col3", [])),
        13: sum(x["monto"] for m in matched for x in m.get("col4", [])),
    }
    ws.cell(row=start_row, column=2, value="TOTAL").font = _font(bold=True)
    for col, total in totals.items():
        t = ws.cell(row=start_row, column=col, value=total)
        t.number_format = _NUM_FMT
        t.font = _font(bold=True)
        t.alignment = _align("right")
        t.fill = _fill(_YELLOW)
    start_row += 1

    _add_table(ws, "CrucesDebitoExtracto",  f"A{header_row}:C{last_data_row}")
    _add_table(ws, "CrucesHaberMayor",      f"D{header_row}:F{last_data_row}")
    _add_table(ws, "CrucesDebeMayor",       f"H{header_row}:J{last_data_row}")
    _add_table(ws, "CrucesCreditoExtracto", f"K{header_row}:M{last_data_row}")

    return start_row + 1


_CATEGORIA_LABELS = {
    "col1": "Débitos no Contabilizados",
    "col2": "No Debitados en Extracto",
    "col3": "No Acreditados",
    "col4": "Créditos no Contabilizados",
}
_VIOLET = "6D28D9"


def _write_ajustes_section(
    ws, start_row: int,
    ajustes: list,
    saldo_original: float,
    saldo_ajustado: float,
) -> int:
    """Ajustes de saldo contable por errores de meses anteriores: registra el
    monto/motivo de cada ajuste y qué pendientes quedaron escritos de baja por él,
    como respaldo de auditoría (por qué el saldo contable no matchea el Mayor crudo)."""

    c = ws.cell(row=start_row, column=1, value="🔧 Ajustes de Saldo Contable (errores de meses anteriores)")
    c.font = _font(bold=True, size=12, color=_DARK_BLUE)
    c.fill = _fill(_PURPLE)
    for col in range(2, 14):
        ws.cell(row=start_row, column=col).fill = _fill(_PURPLE)
    start_row += 1

    ws.cell(row=start_row, column=1, value="Saldo Contable Original (Mayor)").font = _font(bold=True)
    v = ws.cell(row=start_row, column=3, value=saldo_original)
    v.number_format = _NUM_FMT
    v.alignment = _align("right")
    start_row += 1

    ws.cell(row=start_row, column=1, value="Saldo Contable Ajustado").font = _font(bold=True)
    v = ws.cell(row=start_row, column=3, value=saldo_ajustado)
    v.number_format = _NUM_FMT
    v.alignment = _align("right")
    start_row += 2

    for idx, ajuste in enumerate(ajustes, start=1):
        c = ws.cell(row=start_row, column=1, value=f"Ajuste {idx}: {ajuste['motivo'] or '(sin motivo)'}")
        c.font = _font(bold=True, size=10, color=_VIOLET)
        m = ws.cell(row=start_row, column=4, value=ajuste["monto"])
        m.number_format = _NUM_FMT
        m.font = _font(bold=True)
        m.alignment = _align("right")
        start_row += 1

        header_row = start_row
        _set_header_row(ws, start_row, {1: "Fecha", 2: "Concepto", 3: "Categoría", 4: "Monto"}, _VIOLET, size=9)
        start_row += 1

        items = ajuste.get("items") or []
        if not items:
            ws.cell(row=start_row, column=1, value="— Sin ítems —").font = _font(color="888888")
            start_row += 1
        else:
            for i, item in enumerate(items):
                bg = _ORANGE if item.get("mes_anterior") else (_GREY_LIGHT if i % 2 == 0 else _WHITE)
                row_vals = [
                    _fmt_date(item), item["descripcion"],
                    _CATEGORIA_LABELS.get(item.get("categoria"), item.get("categoria")),
                    item["monto"],
                ]
                _write_data_row(ws, start_row, row_vals, bg)
                start_row += 1
            _add_table(ws, f"AjusteItems{idx}", f"A{header_row}:D{start_row - 1}")

        start_row += 2

    return start_row
